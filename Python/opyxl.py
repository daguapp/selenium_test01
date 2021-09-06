from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
class TestOpenpyxl:



    def create_xlsx(self): #实例化workbook对象激活并编辑，最后创建一个xls实体
        #实例化
        self.wb=Workbook()
        #激活worksheet
        self.ws=self.wb.active
        #创建xlsx文件
        self.ws["A1"]="id" #数据分配到虚拟单元格（1
        self.ws["B2"]="1" #数据分配到虚拟单元格（2
        # self.ws.append([2,3,4,5])
        self.ws['A2'] = datetime.datetime.now().strftime("%Y-%m-%d")
        self.wb.save('sample.xlsx')

    def reading_xlsx(self): #读取存在的文件内容
        wb=load_workbook(r"C:\Users\admin\Desktop\选品方舟测试用例\选品方舟V3.0.2.2.xlsx") #使用load_work读取本地文件
        ws=wb.active #打开默认工作簿
        a=ws["A1"].value
        b=ws.cell(row=2,column=3).value #
        print("%s\n%s"%(a,b))

    def write_xlsx(self): #xlsx写入数据
        pass

if __name__ == '__main__':
    TestOpenpyxl().reading_xlsx()