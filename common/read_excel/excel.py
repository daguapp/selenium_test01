import openpyxl

def read_xls(path,sheet_name):
    """读取excel"""
    #加载目录
    el=openpyxl.load_workbook(path)

    #获取表
    ex=el[sheet_name]

    #打印最大行和列
    # print(ex.max_row,ex.max_column)

    sheet_list=[]
    for row in range(2,ex.max_row+1):
        col=[]
        for co in range(1,ex.max_column+1):
            col.append(ex.cell(row,co).value)
        sheet_list.append(col)
    return sheet_list
if __name__ == '__main__':
    pass